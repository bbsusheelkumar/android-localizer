# Localization & Validation Logic (Base Version)
# Technologies Used -> Python 3.6.2, OpenPyXL Libraries, PyCharm IDE
# Developed by B B Susheel Kumar
# bb.susheelkumar@yahoo.com

from openpyxl import load_workbook
import xml.etree.ElementTree as ET
import fnmatch
import os
import subprocess

# Enable (1) / Disable (0) Debugging Logs
dLogs = 0


# ----------------------------------> START of STRING EXCEL SHEET INITIALIZATION Logic <--------------------------------

# Name of String Excel Sheet
stringSheetFile = "StringSheet.xlsx"

# Opening Excel Workbook
stringBook = load_workbook(filename=stringSheetFile, read_only=False)
# Go to "Strings" WorkSheet in the Excel File Opened
stringSheet = stringBook['Strings']
# Read only Column "B" Cells (Cell Address)-> String's, So stringCol contains list of all String's in the Excel Sheet
stringCol = stringSheet['B']
# Read only Row "2" Cells (Cell Address)-> Languages, So langRow contains list of all Languages in the Excel Sheet
langRow = stringSheet[2]

# Extract the data in langRow cell addresses to langList (Stores the list of Languages Supported in String Excel Sheet)
langList = []

# ----------------------------------> END of STRING EXCEL SHEET INITIALIZATION Logic <----------------------------------


# Row index of String Excel Sheet
rowCounter = 0
# Counter which gives the number of String's in XML which are same as in String Excel Sheet
matched = 0

# Counter which gives the number of String's in XML which are found in String Excel Sheet
foundCounter = 0


# -------------------------------------> START of STRING EXCEL SHEET SCANNING Logic <------------------------------------
# Extracting the data in langRow cell addresses and adding it to langList
# to make a list of all the available languages in the given String Excel sheet
for cell in langRow:
    if cell.value is not None:
        langList.append(cell.value.strip())

# Path of the Project/Folder/File to be parsed (Default: Current Working Directory)
my_path = os.getcwd()

# Stores the paths of all strings.xml files in a given Project/Folder/File
strFileList = []
# Stores the language of the respective strings.xml file
# Index of strFileList => Index in strFileLang:: Language of strings.xml file in the path as in the index of strFileList
strFileLang = []

# Contains the Column's related to various languages. (For Korea => Column 'K' which is nothing but 11th Column)
# valueList = ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z',
#              'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ']
valueList = [6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26,
             27, 28, 29, 30, 31, 32, 33, 34, 35, 36]
# -------------------------------------> END of STRINGS EXCEL SHEET SCANNING Logic <--------------------------------------


# -----------------------------------------> START of FILE SCANNING Logic <-------------------------------------------
# Logic to walk recursively through various folders, files in the given path (my_path)
for root, dirs, files in os.walk(my_path):
    if dLogs: print("Root: %s, Dirs: %s, File: %s" % (root, dirs, files))
    valueDir = fnmatch.filter(dirs, "values*")
    for filteredDir in valueDir:
        if dLogs: print(dirs)
        if filteredDir in langList:
            dirFiles = os.listdir(os.path.join(root, filteredDir))
            xmlfiles = fnmatch.filter(dirFiles, "*strings.xml")
            if dLogs: print(xmlfiles)
            count = 0
            while count < len(xmlfiles):
                strFileLang.append(langList.index(filteredDir))
                count += 1
            strFileList.extend(os.path.join(root, filteredDir, f) for f in xmlfiles)
# -----------------------------------------> END of FILE SCANNING Logic <-------------------------------------------


# Prints the list of column number in the Excel Sheet to be checked for the respective strings.xml
if dLogs: print("\n Columns to be checked: "+strFileLang+"\n")

# Counter to iterate through the list of path's having strings.xml
fileCounter = 0


# ---------------------------------------------> START of MAIN Logic <------------------------------------------------
for strFile in strFileList:
    if dLogs: print("File Number: "+fileCounter+" Related Column in String Sheet: "+valueList[strFileLang[fileCounter]])
    if dLogs: print("File Path: "+strFile+" "+"\n")
    # sheetColIdx gives the column in String Excel Sheet to be checked for the respective language
    sheetColIdx = valueList[strFileLang[fileCounter]]
    # Parsing the strings.xml file as an Element Tree (DOM Tree)
    eleTree = ET.parse(strFile)
    # Get the root element of the XML file (root is generally <resources> tag)
    treeRoot = eleTree.getroot()
    # Listing all the <string> elements
    strEle = eleTree.findall('string')

    # Logic to Validate the XML String with String in Excel Sheet
    # Get the String's in XML and check them in the Excel Sheet for respective Values
    # Compare values of respective String's between XML and Excel File
    # If not similar then replace, else skip to next String in strings.xml
    for cell in stringCol:
        # String's start after second row in the String Excel Sheet
        if rowCounter >= 2:
            for stringTag in strEle:
                # Get the String Attribute from <string> tag, it gives String Number
                stringName = stringTag.get('STRID')
                # NULL Check for Cell Value in Excel and string Number in XML
                if cell.value is not None and StringName is not None and stringName.startswith('STRING'):
                    if cell.value.strip() == stringName:
                        if dLogs: print("FOUND : ", stringName)
                        # Get the respective String Value for a String Number from XML
                        stringValue = stringTag.text
                        # While comparing the values in the XML, we need to remove double quotes at the Start and End
                        if stringValue.startswith('"') and stringValue.endswith('"'):
                            stringValue = stringValue[1:-1]
                        if dLogs: print("XML : ", stringValue)
                        # ----------> START: Get the respective String Value for a String Number from EXCEL <---------------
                        sheetValue = stringSheet.cell(row=rowCounter + 1, column=sheetColIdx).value
                        # In case of string value which has Multiple Lines in the Excel Sheet, split them as lines
                        stringLines = sheetValue.splitlines()
                        # Join the lines split as a single line (Multiple Line String's should have '\n' embedded in it
                        sheetValue = "\\n".join(stringLines)
                        # ----------> END: Get the respective String Value for a String Number from EXCEL <---------------
                        if dLogs: print("EXCEL : ", sheetValue)
                        if sheetValue == stringValue:
                            matched += 1
                            if dLogs: print("MATCHED")
                        else:
                            if dLogs: print("NOT MATCHED")
                            # If not matched, replace the value in strings.xml with Excel Sheet Value.
                            stringTag.text = '"'+sheetValue+'"'
                            if dLogs: print("CHANGED in XML")
                        if dLogs: print("\n")
                        foundCounter += 1
        rowCounter += 1
    fileCounter += 1
    rowCounter = 0
    # Write and Save the strings.xml file
    eleTree.write(strFile)
# ---------------------------------------------> END of MAIN Logic <------------------------------------------------


# ---------------------------------------------> START of RESULTS <------------------------------------------------
print("\nString's FOUND : %d" % foundCounter)
print("String's MATCHED : %d\nString's NOT MATCHED : %d" % (matched, (foundCounter - matched)))
# ----------------------------------------------> END of RESULTS <-------------------------------------------------
